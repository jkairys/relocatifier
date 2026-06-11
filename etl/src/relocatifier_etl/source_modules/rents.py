"""Median weekly house rents: QLD RTA suburb medians + NSW bond lodgements.

QLD: the RTA publishes median weekly rents for new tenancies per quarter,
by suburb and dwelling type, in one consolidated workbook (sheet
"4 sub-rents"). We prefer the 3-bedroom house series ("House 3") — the RTA
has no all-houses series, so where House 3 is suppressed we fall back to
House 4 then House 2, and we never use "All dwellings" (it mixes in flats,
which would corrupt the derived gross yield against house prices). The RTA
suppresses small-sample quarters, so per series we take the latest published
quarter within the last four (series preference beats recency: a year-old
3-bed median is closer to "3-bed house rent" than this quarter's 4-bed one).
Suburb names map to SAL via ctx.name_lookup, exact match only (ADR-0001);
unmatched counts are reported.

The RTA suppresses thin markets at suburb level (only ~500 of 3,200 QLD SALs
get a suburb median), but the same workbook carries a POSTCODE-level sheet
("1 pc-rents", identical layout) with much broader coverage. QLD SALs with
no suburb-sheet house rent fall back to the house median (same series
preference and lookback) of their dominant-overlap POA 2021 postcode —
suburb-sheet values always win over postcode values, and the fallback is
flagged as approximated in the metric notes per ADR-0001.

NSW: Fair Trading publishes individual rental bond lodgements monthly at
POSTCODE level (no suburb). We pool the most recent 12 monthly files, keep
dwelling type "H" (house), and take the median weekly rent per postcode
(suppressed below MIN_NSW_BONDS lodgements). Postcode -> SAL uses the ABS
POA 2021 boundaries: each NSW SAL is assigned the postcode with the largest
area overlap (EPSG:3577 for area math), so the metric is an approximation
for suburbs that straddle postcodes (flagged in the metric notes per
ADR-0001).
"""

from __future__ import annotations

from statistics import median
from typing import Iterable, Iterator

import geopandas as gpd
import openpyxl

from ..context import BuildContext
from ..sources import Source

# RTA median rents quarterly data — consolidated bond-statistics workbook,
# updated in place each quarter (URL stable since April 2023).
# Source page: https://www.rta.qld.gov.au/forms-resources/rta-quarterly-data/median-rents-quarterly-data
# Licence: CC BY 4.0.
RTA_RENTS = Source(
    name="QLD RTA median rents quarterly data (bond statistics workbook)",
    url="https://www.rta.qld.gov.au/sites/default/files/2023-04/rta-bond-statistics.xlsx",
    filename="rta-bond-statistics.xlsx",
)

# ABS Postal Areas (POA) 2021 digital boundaries, GDA2020 shapefile — same
# ABS digital-boundary-files page as the SAL backbone in sources.py.
# Source page: https://www.abs.gov.au/statistics/standards/australian-statistical-geography-standard-asgs/edition-3-july-2021-june-2026/access-and-downloads/digital-boundary-files
# Licence: CC BY 4.0.
POA_BOUNDARIES = Source(
    name="ABS POA 2021 boundaries (GDA2020 shapefile)",
    url=(
        "https://www.abs.gov.au/statistics/standards/"
        "australian-statistical-geography-standard-asgs/"
        "edition-3-july-2021-june-2026/access-and-downloads/"
        "digital-boundary-files/POA_2021_AUST_GDA2020_SHP.zip"
    ),
    filename="POA_2021_AUST_GDA2020_SHP.zip",
)

# NSW Fair Trading rental bond lodgements, one workbook per month, the most
# recent 12 months. Listing page (new files appear here monthly, with
# irregular file names — extend/rotate this dict when refreshing):
# https://www.nsw.gov.au/housing-and-construction/rental-forms-surveys-and-data/rental-bond-data
# Dataset record: https://data.nsw.gov.au/data/dataset/rental-bond-lodgement
_NSW_BOND_URLS = {
    "2025-05": "2025-06/rental-bond-lodgement-data-may-2025.xlsx",
    "2025-06": "2025-08/rentalbond_lodgements_june_2025_0.xlsx",
    "2025-07": "2025-08/rental-bond-lodgement-data-july-2025.xlsx",
    "2025-08": "2025-09/rentalbond_lodgements_august_2025.xlsx",
    "2025-09": "2025-10/rentalbond_lodgements_september25.xlsx",
    "2025-10": "2025-11/rentalbond_lodgements_october_2025.xlsx",
    "2025-11": "2025-12/rentalbond_lodgements_november_2025.xlsx",
    "2025-12": "2026-01/rentalbond_lodgements_december_2025.xlsx",
    "2026-01": "2026-02/rentalbond_lodgements_january_2026.xlsx",
    "2026-02": "2026-03/rentalbond_lodgements_february_2026.xlsx",
    "2026-03": "2026-04/rentalbond_lodgements_march_2026.xlsx",
    "2026-04": "2026-05/rentalbond_lodgements_april_2026.xlsx",
}

NSW_BOND_SOURCES = [
    Source(
        name=f"NSW rental bond lodgements {month}",
        url=f"https://www.nsw.gov.au/sites/default/files/noindex/{path}",
        filename=f"nsw_rentalbond_lodgements_{month}.xlsx",
    )
    for month, path in _NSW_BOND_URLS.items()
]

RAW_SOURCES = [RTA_RENTS, POA_BOUNDARIES, *NSW_BOND_SOURCES]

METRICS = {
    "median_rent_house": {
        "label": "Median weekly rent (houses)",
        "format": "aud_per_week",
        "direction": "lower_better",
        "notes": (
            "QLD: RTA median rent for new tenancies, 3-bedroom houses "
            "(falling back to 4- then 2-bedroom where the 3-bed series is "
            "suppressed), latest published quarter within the past year. "
            "Where the RTA suppresses the suburb entirely, the suburb "
            "inherits the postcode-level median (same series rules) of the "
            "postcode with the largest area overlap — approximated, as for "
            "NSW below. "
            "NSW: approximated — median weekly rent of house bond "
            "lodgements over the last 12 months for the postcode with the "
            "largest area overlap with the suburb; suburbs straddling "
            "postcodes inherit their dominant postcode's median."
        ),
    },
}

VINTAGES = {"rents_qld": "2026-Q1", "rents_nsw": "12m to 2026-04"}

# Dwelling-series preference for the QLD "house" rent. The RTA has no
# all-houses series, so 3-bed is canonical and other house sizes are the
# fallback (never townhouses/flats/"All dwellings").
HOUSE_SERIES_PREFERENCE = ("House 3", "House 4", "House 2")

# How many recent quarters a suppressed RTA series may be backfilled from.
RTA_LOOKBACK_QUARTERS = 4

# NSW postcode medians from fewer lodgements than this are suppressed, in the
# spirit of the RTA's own small-count suppression.
MIN_NSW_BONDS = 5

_QUARTER_BY_MONTH = {"Mar": 1, "Jun": 2, "Sep": 3, "Dec": 4}

_RTA_SUBURB_SHEET = "4 sub-rents"
_RTA_POSTCODE_SHEET = "1 pc-rents"
_NSW_HEADER = ("Lodgement Date", "Postcode", "Dwelling Type", "Bedrooms", "Weekly Rent")


# --- pure parsing logic (unit-tested without the raw downloads) -------------


def _cell_number(value) -> float | None:
    """RTA medians: numbers are data; ''/None/anything else is suppressed."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return None


def _suburb_key(value) -> str | None:
    """Suburb-sheet entity cells: any non-empty string is a suburb name."""
    return value.strip() if isinstance(value, str) else None


def _postcode_key(value) -> str | None:
    """Postcode-sheet entity cells: numbers (or digit strings) zero-padded."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"{int(value):04d}"
    if isinstance(value, str) and value.strip().isdigit():
        return f"{int(value.strip()):04d}"
    return None


def _parse_rta_rents(
    rows: Iterable[tuple], entity_header: str, entity_key, lookback: int
) -> tuple[str, dict[str, dict[str, list[float | None]]]]:
    """Parse an RTA median-rents sheet (suburb and postcode sheets share the
    layout, differing only in the entity column's header and cell type).

    Returns (latest_quarter_label, {entity: {dwelling_series: values}}) where
    values are the last `lookback` quarters' medians oldest-first (None =
    suppressed). Row tuples come from openpyxl values_only iteration; layout
    (located by header text, not fixed positions): an entity/"Dwelling"
    header row, then a month row and a year row over the quarter columns,
    then data rows.
    """
    rows = iter(rows)
    entity_col = dwelling_col = None
    for row in rows:
        cells = [c.strip() if isinstance(c, str) else c for c in row]
        if entity_header in cells and "Dwelling" in cells:
            entity_col = cells.index(entity_header)
            dwelling_col = cells.index("Dwelling")
            break
    if entity_col is None:
        raise ValueError(
            f"RTA {entity_header.lower()} sheet: "
            f"no {entity_header}/Dwelling header row found"
        )

    months = next(rows, ())
    years = next(rows, ())
    quarter_cols = [
        i
        for i in range(dwelling_col + 1, len(years))
        if i < len(months)
        and months[i] in _QUARTER_BY_MONTH
        and isinstance(years[i], int)
    ]
    if not quarter_cols:
        raise ValueError(f"RTA {entity_header.lower()} sheet: no quarter columns found")
    recent = quarter_cols[-lookback:]
    latest = recent[-1]
    quarter_label = f"{years[latest]}-Q{_QUARTER_BY_MONTH[months[latest]]}"

    entities: dict[str, dict[str, list[float | None]]] = {}
    for row in rows:
        entity = entity_key(row[entity_col]) if len(row) > entity_col else None
        dwelling = row[dwelling_col] if len(row) > dwelling_col else None
        if entity is None or not isinstance(dwelling, str):
            continue
        values = [_cell_number(row[i]) if len(row) > i else None for i in recent]
        entities.setdefault(entity, {})[dwelling.strip()] = values
    return quarter_label, entities


def parse_rta_suburb_rents(
    rows: Iterable[tuple], lookback: int = RTA_LOOKBACK_QUARTERS
) -> tuple[str, dict[str, dict[str, list[float | None]]]]:
    """Parse the RTA "4 sub-rents" sheet rows: {suburb: {series: values}}."""
    return _parse_rta_rents(rows, "Suburb", _suburb_key, lookback)


def parse_rta_postcode_rents(
    rows: Iterable[tuple], lookback: int = RTA_LOOKBACK_QUARTERS
) -> tuple[str, dict[str, dict[str, list[float | None]]]]:
    """Parse the RTA "1 pc-rents" sheet rows: {postcode: {series: values}}.

    Postcodes arrive as numbers in the sheet and are returned as zero-padded
    4-digit strings, matching the POA boundary codes.
    """
    return _parse_rta_rents(rows, "Postcode", _postcode_key, lookback)


def pick_house_rent(series: dict[str, list[float | None]]) -> float | None:
    """The suburb's house rent from per-series recent-quarter values.

    Series preference beats recency: the most recent non-suppressed House 3
    median wins over any House 4/House 2 value (values are oldest-first).
    """
    for dwelling in HOUSE_SERIES_PREFERENCE:
        for value in reversed(series.get(dwelling, [])):
            if value is not None:
                return value
    return None


def fill_missing_from_postcodes(
    values: dict[str, float],
    sal_to_postcode: dict[str, str],
    postcode_rents: dict[str, float | None],
) -> dict[str, float]:
    """Postcode-derived rents for SALs that lack a suburb-sheet value.

    Suburb-sheet values always win: SALs already present in `values` are never
    overridden. A SAL whose dominant postcode has no house median gets nothing
    (it does not inherit some lesser-overlap postcode's value).
    """
    return {
        sal_code: rent
        for sal_code, postcode in sal_to_postcode.items()
        if sal_code not in values
        and (rent := postcode_rents.get(postcode)) is not None
    }


def parse_nsw_bond_rows(rows: Iterable[tuple]) -> Iterator[tuple[str, float]]:
    """Yield (postcode, weekly_rent) for HOUSE bond lodgements.

    Rows are openpyxl values_only tuples (Lodgement Date, Postcode,
    Dwelling Type, Bedrooms, Weekly Rent); preamble before the header row is
    skipped. Dwelling type "H" = house; rent and postcode arrive as either
    numbers or strings ("U" = unknown) and non-positive rents are dropped.
    """
    rows = iter(rows)
    for row in rows:
        if tuple(c.strip() if isinstance(c, str) else c for c in row[:5]) == _NSW_HEADER:
            break
    else:
        raise ValueError("NSW bond workbook: header row not found")

    for row in rows:
        if len(row) < 5 or row[2] != "H":
            continue
        postcode, rent = row[1], row[4]
        if isinstance(postcode, (int, float)) and not isinstance(postcode, bool):
            postcode = f"{int(postcode):04d}"
        elif isinstance(postcode, str) and postcode.strip().isdigit():
            postcode = f"{int(postcode.strip()):04d}"
        else:
            continue
        if isinstance(rent, str):
            rent = rent.replace(",", "").strip()
            if not rent.replace(".", "", 1).isdigit():
                continue
            rent = float(rent)
        elif not isinstance(rent, (int, float)) or isinstance(rent, bool):
            continue
        if rent > 0:
            yield postcode, float(rent)


def postcode_medians(
    lodgements: Iterable[tuple[str, float]], min_count: int = MIN_NSW_BONDS
) -> dict[str, float]:
    """Median weekly rent per postcode, suppressing small samples."""
    by_postcode: dict[str, list[float]] = {}
    for postcode, rent in lodgements:
        by_postcode.setdefault(postcode, []).append(rent)
    return {
        postcode: float(median(rents))
        for postcode, rents in by_postcode.items()
        if len(rents) >= min_count
    }


def dominant_zone_by_area(
    targets: gpd.GeoDataFrame, zones: gpd.GeoDataFrame, zone_col: str
) -> dict[str, str]:
    """Assign each target (sal_code) the zone with the largest area overlap.

    Both frames must already share a planar CRS (we use EPSG:3577 in build so
    intersection areas are in square metres). Targets with no overlap at all
    are simply absent from the result.
    """
    pieces = gpd.overlay(
        targets[["sal_code", "geometry"]],
        zones[[zone_col, "geometry"]],
        how="intersection",
        keep_geom_type=True,
    )
    if pieces.empty:
        return {}
    pieces["overlap_area"] = pieces.geometry.area
    best = pieces.loc[pieces.groupby("sal_code")["overlap_area"].idxmax()]
    return dict(zip(best["sal_code"], best[zone_col]))


# --- build -------------------------------------------------------------------


def _require(ctx: BuildContext, source: Source):
    path = ctx.raw_dir / source.filename
    if not path.exists():
        raise SystemExit(f"missing {path} — run `etl fetch` first")
    return path


def _load_poa(ctx: BuildContext, prefix: str) -> gpd.GeoDataFrame:
    """POA 2021 polygons whose postcode starts with `prefix`, as a frame with
    columns postcode, geometry (POA covers geographic postcodes only)."""
    poa = gpd.read_file(f"zip://{_require(ctx, POA_BOUNDARIES)}")
    poa = poa[poa["POA_CODE21"].str.startswith(prefix) & poa.geometry.notna()].copy()
    return poa.rename(columns={"POA_CODE21": "postcode"})


def _build_qld(ctx: BuildContext) -> dict[str, float]:
    wb = openpyxl.load_workbook(_require(ctx, RTA_RENTS), read_only=True)
    try:
        quarter, suburbs = parse_rta_suburb_rents(
            wb[_RTA_SUBURB_SHEET].iter_rows(values_only=True)
        )
        pc_quarter, postcodes = parse_rta_postcode_rents(
            wb[_RTA_POSTCODE_SHEET].iter_rows(values_only=True)
        )
    finally:
        wb.close()
    if quarter != VINTAGES["rents_qld"]:
        raise SystemExit(
            f"RTA workbook latest quarter is {quarter} but VINTAGES says "
            f"{VINTAGES['rents_qld']} — update rents.VINTAGES"
        )
    if pc_quarter != quarter:
        raise SystemExit(
            f"RTA postcode sheet latest quarter is {pc_quarter} but the "
            f"suburb sheet says {quarter} — workbook is inconsistent"
        )

    values: dict[str, float] = {}
    matched = unmatched = no_series = 0
    for suburb, series in suburbs.items():
        rent = pick_house_rent(series)
        if rent is None:
            no_series += 1
            continue
        sal_code = ctx.name_lookup.sal_code(suburb, "QLD")
        if sal_code is None:
            unmatched += 1
            continue
        matched += 1
        values[sal_code] = rent
    print(
        f"[rents] QLD {quarter}: {matched} suburbs matched to SAL, "
        f"{unmatched} unmatched by name, {no_series} without a house series"
    )

    # Postcode fallback for RTA-suppressed suburbs: each QLD SAL still missing
    # a rent gets its dominant-overlap postcode's house median. "4xxx" covers
    # all geographic QLD postcodes (QLD's 9xxx ranges are non-geographic and
    # absent from POA), so border SALs compete among QLD postcodes only —
    # consistent with the RTA data being QLD-only.
    postcode_rents = {pc: pick_house_rent(series) for pc, series in postcodes.items()}
    qld = ctx.suburbs[ctx.suburbs["state"] == "QLD"]
    missing = qld[~qld["sal_code"].isin(values)]
    sal_to_postcode = dominant_zone_by_area(
        missing.to_crs(epsg=3577), _load_poa(ctx, "4").to_crs(epsg=3577), "postcode"
    )
    fallback = fill_missing_from_postcodes(values, sal_to_postcode, postcode_rents)
    values.update(fallback)
    print(
        f"[rents] QLD postcode fallback: {len(missing)} suburbs without a suburb "
        f"median, {len(sal_to_postcode)} assigned a dominant postcode, "
        f"{len(fallback)} filled from postcode medians"
    )
    return values


def _build_nsw(ctx: BuildContext) -> dict[str, float]:
    lodgements: list[tuple[str, float]] = []
    for source in NSW_BOND_SOURCES:
        wb = openpyxl.load_workbook(_require(ctx, source), read_only=True)
        try:
            data_sheet = next(s for s in wb.sheetnames if s != "Definitions")
            lodgements.extend(parse_nsw_bond_rows(wb[data_sheet].iter_rows(values_only=True)))
        finally:
            wb.close()
    medians = postcode_medians(lodgements)
    print(
        f"[rents] NSW: {len(lodgements)} house bonds over {len(NSW_BOND_SOURCES)} months, "
        f"{len(medians)} postcodes with >= {MIN_NSW_BONDS} bonds"
    )

    # "2xxx" covers NSW (plus ACT, which loses the area-overlap contest
    # outside the border fringe).
    poa = _load_poa(ctx, "2")

    nsw = ctx.suburbs[ctx.suburbs["state"] == "NSW"]
    sal_to_postcode = dominant_zone_by_area(
        nsw.to_crs(epsg=3577), poa.to_crs(epsg=3577), "postcode"
    )

    values: dict[str, float] = {}
    no_median = 0
    for sal_code, postcode in sal_to_postcode.items():
        rent = medians.get(postcode)
        if rent is None:
            no_median += 1
            continue
        values[sal_code] = rent
    print(
        f"[rents] NSW: {len(sal_to_postcode)}/{len(nsw)} suburbs assigned a dominant "
        f"postcode, {len(values)} with a rent median, {no_median} postcode had no median"
    )
    return values


def build(ctx: BuildContext) -> dict[str, dict[str, float | None]]:
    values = {**_build_qld(ctx), **_build_nsw(ctx)}
    return {
        sal_code: {"median_rent_house": rent}
        for sal_code, rent in values.items()
        if sal_code in ctx.sal_codes
    }
